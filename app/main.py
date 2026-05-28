from flask import Flask, render_template, request, redirect, url_for, send_file
import os
import pandas as pd
import json
import io
from datetime import datetime
from urllib.parse import unquote
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


import base64
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image
import tempfile


from auditoria.validaciones import preparar_datos
from auditoria.reglas import cargar_reglas, ejecutar_reglas
from auditoria.anomalias_ia import detectar_anomalias
from auditoria.anomalias_ia import ejecutar_analisis_completo
from auditoria.anomalias_ia import generar_insights_completo
from auditoria.anomalias_ia import analizar_inconsistencias


app = Flask(__name__)

# ===============================
# CONFIGURACIÓN (RUTAS ABSOLUTAS)
# ===============================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "data", "uploads")
BASES_FOLDER = os.path.join(BASE_DIR, "data", "archivos_base")
METADATA_FILE = os.path.join(BASE_DIR, "data", "metadata.json")
RESULTADOS_FILE = os.path.join(BASE_DIR, "data", "resultados.json")
RESULTADOS_IA = os.path.join(BASE_DIR, "data", "resultados_ia.json")
BASES_METADATA = os.path.join(BASES_FOLDER, "bases_metadata.json")
HISTORICO_FILE = os.path.join(BASE_DIR, "data", "historico_auditorias.json" )
ALERTAS_FILE = os.path.join(BASE_DIR, "data", "alertas.json")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(BASES_FOLDER, exist_ok=True)

# ===============================
# CONFIGURACIÓN DE BASES Y AUDITORÍAS
# ===============================

BASE_CONTRATOS = {
    "Inventario": ["owner"],
    "Forecast": ["Product", "Forecast"],
    "Status": ["Product", "Status"],
    "LeadTime": ["product"],
    "Activations":["item", "date_active/inactive"]
}

AUDITORIAS = {
  
    "leadtimes": {
        "nombre": "Validación leadtimes",
        "bases_requeridas": ["Inventario"],
        "reglas": ["LT_01", "LT_02", "LT_03"]
    },
    "leadtime_vs_base": {
        "nombre": "Validación leadtimes vs base",
        "bases_requeridas": ["LeadTime", "Inventario"],
        "reglas": ["LT_04"]
    },
    "lifecycle_status": {
        "nombre": "Validación Lifecycle",
        "bases_requeridas": ["Activations"],
        "reglas": ["LCS_TIME"]
    },
    "datos_faltantes":{
        "nombre": "Validación datos faltantes",
        "bases_requeridas":[],
        "reglas": ["MS_01"]
    }
}


# ===============================
# FUNCIONES AUXILIARES
# ===============================

def cargar_json(path):
    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)
    
def calcular_score(df, inconsistencias, anomalias ):
    total= len(df) if len (df) > 0 else 1
    error_rate=len(inconsistencias) / total
    anomaly_rate = len (anomalias) / total

    score = 200 - (error_rate * 60 + anomaly_rate * 40) * 100
    return round (max(score, 0), 2)
    
def generar_alertas(df, inconsistencias, anomalias, estadisticas, analisis_inc):
    alertas = []
    total = len(df) if len (df) > 0 else 1

    ratio_inc = len (inconsistencias) / total
    if ratio_inc > 0.1:
        alertas.append({
            "tipo": "Calidad",
            "nivel":"Critica",
            "mensaje": f"Mas del 10% de inconsistencias({round(ratio_inc*100, 2)}%)"
        })

    ratio_ano = len (anomalias) / total
    if ratio_ano > 0.08:
        alertas.append({
            "tipo": "Anomalias",
            "nivel":"Alta",
            "mensaje": f"Alta concentración de anomalias({round(ratio_inc*100, 2)}%)"
        })


    for c in estadisticas.get("calidad_datos", []):
        if c["porcentaje_nulos"] > 20:
            alertas.append({
                            "tipo": "Datos",
                            "nivel": "Alta",
                            "mensaje": f"{c['columna']} tiene {c['porcentaje_nulos']}% nulos"
                        })

    return alertas





def guardar_json(path, data):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def cargar_metadata():
    return cargar_json(METADATA_FILE)

def guardar_metadata(data):
    guardar_json(METADATA_FILE, data)

def cargar_bases_metadata():
    return cargar_json(BASES_METADATA)

def guardar_bases_metadata(data):
    guardar_json(BASES_METADATA, data)

def guardar_resultados(data):
    guardar_json(RESULTADOS_FILE, data)

def cargar_resultados():
    return cargar_json(RESULTADOS_FILE)

def guardar_resultados_ia(data):
    guardar_json(RESULTADOS_IA, data)

def cargar_resultados_ia():
    return cargar_json(RESULTADOS_IA)


# ===============================
# VISTA DASHBOARD
# ===============================


@app.route("/", methods=["GET", "POST"])
def index():

    inconsistencias=[]

    severidad_count = {}
    regla_count = {}

    resumen = {
        "archivo": "-",
        "registros_auditados": 0,
        "inconsistencias": 0,
        "anomalias": 0,
        "fecha": "-"
    }

    resumen_path = os.path.join(BASE_DIR, "data", "resumen_auditoria.json")

    if os.path.exists(resumen_path) and os.path.getsize(resumen_path) > 0:
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)

    # ===============================
    # 🔹 EJECUTAR AUDITORÍA (POST)
    # ===============================
    if request.method == "POST":

        tipo_auditoria = request.form.get("auditoria_tipo")

        if not tipo_auditoria:
            return render_template("index.html", resumen=resumen,
                                   error="Debe seleccionar un tipo de auditoría")

        config = AUDITORIAS.get(tipo_auditoria)

        if not config:
            return render_template("index.html", resumen=resumen,
                                   error=f"Auditoría no válida: {tipo_auditoria}")

        bases = cargar_bases_metadata()

        tipos_cargados = [b["tipo"].lower() for b in bases]

        faltantes = [
            b for b in config["bases_requeridas"]
            if b.lower() not in tipos_cargados
        ]

        if faltantes:
            return render_template("index.html", resumen=resumen,
                                   error=f"Faltan archivos base requeridos: {faltantes}")

        archivo = request.files.get("archivo")

        if not archivo or not archivo.filename:
            return render_template("index.html", resumen=resumen,
                                   error="Debe cargar un archivo")

        filename = archivo.filename
        ruta = os.path.join(UPLOAD_FOLDER, filename)
        archivo.save(ruta)

        resultado = preparar_datos(ruta)
        df = resultado["dataframe"]

        # Metadata
        metadata = cargar_metadata()
        metadata.append({
            "filename": filename,
            "rows": len(df),
            "columns": len(df.columns),
            "uploaded_at": datetime.now().strftime("%m-%d-%Y %H:%M")
        })
        guardar_metadata(metadata)

        # Bases
        bases_data = {}
        for b in bases:
            if b["tipo"].lower() in [x.lower() for x in config["bases_requeridas"]]:
                bases_data[b["tipo"]] = os.path.join(BASES_FOLDER, b["filename"])

        reglas_activas = config["reglas"]

        inconsistencias = ejecutar_reglas(df, bases_data, reglas_activas, tipo_auditoria)
        guardar_resultados(inconsistencias)

        
        analisis = ejecutar_analisis_completo(df, inconsistencias)
        anomalias = analisis["anomalias"]
        
        score = calcular_score(df, inconsistencias, anomalias)

        historico = cargar_json(HISTORICO_FILE)

        historico.append({
            "archivo": filename,
            "fecha": datetime.now().strftime("%m-%d-%Y %H:%M"),
            "registros": len(df),
            "auditoria": tipo_auditoria,
            "inconsistencias": len(inconsistencias),    
            "anomalias": len(anomalias),
            "score_calidad": score,
            "tipo_auditoria": tipo_auditoria
        })

        guardar_json(HISTORICO_FILE, historico)

        
        alertas = generar_alertas(
            df,
            inconsistencias,
            anomalias,
            analisis["estadisticas"],
            analisis["analisis_inconsistencias"]
        )

        guardar_json(ALERTAS_FILE, alertas)


        
        analisis_inc = analizar_inconsistencias(inconsistencias)

        guardar_json(
            os.path.join(BASE_DIR, "data", "analisis_inconsistencias.json"),
            analisis_inc
        )


        anomalias = analisis ["anomalias"]

        guardar_resultados_ia(analisis["anomalias"])
        guardar_json(os.path.join(BASE_DIR, "data", "estadisticas.json"), analisis["estadisticas"])
        guardar_json(os.path.join(BASE_DIR, "data", "insights.json"), analisis["insights"])


        resumen = {
            "archivo": filename,
            "registros_auditados": len(df),
            "inconsistencias": len(inconsistencias),
            "anomalias": len(anomalias),
            "fecha": datetime.now().strftime("%m-%d-%Y %H:%M"),
            "tipo_auditoria": tipo_auditoria
        }

        guardar_json(resumen_path, resumen)


            



    # ===============================
    # 🔹 SIEMPRE CARGAR RESULTADOS
    # ===============================
    inconsistencias = cargar_resultados()

    # ===============================
    # 🔹 SEVERIDAD / REGLAS
    # ===============================
    if inconsistencias:
        for inc in inconsistencias:
            sev = inc.get("severidad", "Desconocido")
            reg = inc.get("regla_id", "Sin regla")

            severidad_count[sev] = severidad_count.get(sev, 0) + 1
            regla_count[reg] = regla_count.get(reg, 0) + 1
    else:
        severidad_count = {"Sin datos": 1}
        regla_count = {"Sin datos": 1}

    # ===============================
    # 🔹 LIFECYCLE
    # ===============================
    status_count = {}
    expected_count = {}

    for inc in inconsistencias:
        status_actual = inc.get("status_actual", "Sin status")
        status_count[status_actual] = status_count.get(status_actual, 0) + 1

        status_esperado = inc.get("status_esperado", "Sin esperado")
        expected_count[status_esperado] = expected_count.get(status_esperado, 0) + 1

    # ===============================
    # 🔹 OWNER (SLICER)
    # ===============================
  
    owner_detalle = {}

    for inc in inconsistencias:

        owner = inc.get("owner", "SIN OWNER")
        regla = inc.get("regla_id", "SIN REGLA")
        severidad = inc.get("severidad", "SIN SEVERIDAD")

        if not owner or owner == "None":
            owner = "SIN OWNER"

        if owner not in owner_detalle:
            owner_detalle[owner] = {
                "reglas": {},
                "severidad": {}
            }


        owner_detalle[owner]["reglas"][regla] = \
            owner_detalle[owner]["reglas"].get(regla, 0) + 1

        owner_detalle[owner]["severidad"][severidad] = \
            owner_detalle[owner]["severidad"].get(severidad, 0) + 1


    # ===============================
    # 🔹 RENDER
    # ===============================
    return render_template(
        "index.html",
        resumen=resumen,
        severidad_count=severidad_count,
        regla_count=regla_count,
        status_count=status_count,
        expected_count=expected_count,
        owner_data=json.dumps(owner_detalle)
    )



# ===============================
# VISTA DATOS
# ===============================

@app.route("/datos")
def datos():
    return render_template(
        "datos.html",
        archivos=cargar_metadata(),
        bases=cargar_bases_metadata()
    )


# ===============================
# RESULTADOS / ANÁLISIS
# ===============================


@app.route("/resultados")
def resultados():

    resumen_path = os.path.join(BASE_DIR, "data", "resumen_auditoria.json")

    auditoria_tipo = ""

    if os.path.exists(resumen_path):
        with open(resumen_path, "r", encoding="utf-8") as f:
            resumen = json.load(f)
            auditoria_tipo = resumen.get("tipo_auditoria", "")

    return render_template(
        "resultados.html",
        inconsistencias=cargar_resultados(),
        auditoria_tipo=auditoria_tipo 
    )


@app.route("/anomalias")
def anomalias():
    

    estadisticas_path = os.path.join(BASE_DIR, "data", "estadisticas.json")
    analisis_inc_path = os.path.join(BASE_DIR, "data", "analisis_inconsistencias.json")
    insights_path = os.path.join(BASE_DIR, "data", "insights.json")
    alertas_path = ALERTAS_FILE

    estadisticas = {}
    analisis_inc = {}
    insights = []
    alertas = []

    if os.path.exists(estadisticas_path):
        estadisticas = cargar_json(estadisticas_path)

    if os.path.exists(analisis_inc_path):
        analisis_inc = cargar_json(analisis_inc_path)

    if os.path.exists(insights_path):
        insights = cargar_json(insights_path)

    if os.path.exists(alertas_path):
        alertas = cargar_json(alertas_path)

    return render_template(
        "anomalias.html",
        anomalias=cargar_resultados_ia(),
        estadisticas=estadisticas,
        insights=insights,
        alertas=alertas
    )



# ===============================
# VER ARCHIVO
# ===============================

@app.route("/datos/<filename>")
def ver_datos(filename):
    ruta = os.path.join(UPLOAD_FOLDER, filename)

    resultado = preparar_datos(ruta)
    df = resultado["dataframe"]

    preview = df.head(50)

    return render_template(
        "ver_datos.html",
        filename=filename,
        columnas=df.columns.tolist(),
        filas=preview.values.tolist()
    )


# ===============================
# ELIMINAR ARCHIVO
# ===============================

@app.route("/datos/eliminar/<filename>", methods=["POST"])
def eliminar_dato(filename):
    ruta = os.path.join(UPLOAD_FOLDER, filename)

    if os.path.exists(ruta):
        os.remove(ruta)

    metadata = cargar_metadata()
    metadata = [m for m in metadata if m["filename"] != filename]
    guardar_metadata(metadata)

    return redirect(url_for("datos"))


# ===============================
# ARCHIVOS BASE
# ===============================

@app.route("/archivos_base/upload", methods=["POST"])
def upload_base():
    archivo = request.files.get("archivo")
    tipo = request.form.get("tipo")

    if not archivo or not tipo:
        return redirect(url_for("datos"))

    ruta = os.path.join(BASES_FOLDER, archivo.filename)
    archivo.save(ruta)

    df_dict = pd.read_excel(ruta, sheet_name=None)

    dfs_validos = []

    for name, df_sheet in df_dict.items():
        cols = (
            df_sheet.columns
            .astype(str)
            .str.strip()
            .str.lower()
            .str.replace(" ", "", regex=False)
            .str.replace("-", "", regex=False)
        )

        
        if tipo == "Activations":
            if any("item" in col for col in cols):
                dfs_validos.append(df_sheet)

        elif tipo == "Inventario":
            if any("owner" in col for col in cols):
                dfs_validos.append(df_sheet)

        else:
            if any("product" in col for col in cols):
                dfs_validos.append(df_sheet)


    if not dfs_validos:
        return render_template(
            "datos.html",
            archivos=cargar_metadata(),
            bases=cargar_bases_metadata(),
            error="El archivo no contiene ninguna hoja con Product"
        )

    df = pd.concat(dfs_validos, ignore_index=True)

    columnas_requeridas = BASE_CONTRATOS.get(tipo, [])

    cols_df = [
        str(c).strip().lower().replace(" ", "").replace("-", "").replace(".", "")
        for c in df.columns
    ]

    cols_req = [
        str(c).strip().lower().replace(" ", "").replace("-", "").replace(".", "")
        for c in columnas_requeridas
    ]

    faltantes = [c for c in cols_req if c not in cols_df]

    if faltantes:
        return render_template(
            "datos.html",
            archivos=cargar_metadata(),
            bases=cargar_bases_metadata(),
            error=f"Faltan columnas requeridas: {faltantes}"
        )

    metadata = cargar_bases_metadata()
    metadata.append({
        "filename": archivo.filename,
        "tipo": tipo,
       
        "uploaded_at": datetime.now().strftime("%m-%d-%Y %H:%M"),
        "columnas": df.columns.tolist()
    })
    guardar_bases_metadata(metadata)

    return redirect(url_for("datos"))


@app.route("/archivos_base/ver/<filename>")
def ver_base(filename):
    filename = unquote(filename)
    ruta = os.path.join(BASES_FOLDER, filename)

    resultado = preparar_datos(ruta)
    df = resultado["dataframe"].head(30)

    return render_template(
        "ver_base.html",
        filename=filename,
        columnas=df.columns.tolist(),
        filas=df.values.tolist()
    )


@app.route("/archivos_base/eliminar/<filename>", methods=["POST"])
def eliminar_base(filename):
    filename = unquote(filename)
    ruta = os.path.join(BASES_FOLDER, filename)

    if os.path.exists(ruta):
        os.remove(ruta)

    metadata = cargar_bases_metadata()
    metadata = [m for m in metadata if m["filename"] != filename]
    guardar_bases_metadata(metadata)

    return redirect(url_for("datos"))




@app.route("/historico/eliminar/<filename>", methods=["POST"])
def eliminar_por_archivo_route(filename):

    historico = cargar_json(HISTORICO_FILE)

   
    historico_filtrado = [
        h for h in historico if h.get("archivo") != filename
    ]

    guardar_json(HISTORICO_FILE, historico_filtrado)

    return redirect(url_for("historico"))





# ===============================
# REGLAS
# ===============================

@app.route("/reglas")
def reglas():
    reglas_config = cargar_reglas()
    return render_template("reglas.html", reglas=reglas_config)


# ===============================
# VER FILA CON ERROR
# ===============================

@app.route("/resultados/fila/<int:fila>")
def ver_fila_error(fila):

    resumen_path = os.path.join(BASE_DIR, "data", "resumen_auditoria.json")

    if not os.path.exists(resumen_path):
        return "No hay auditoría ejecutada", 404

    with open(resumen_path, "r", encoding="utf-8") as f:
        resumen = json.load(f)

    filename = resumen.get("archivo")

    if not filename:
        return "No se pudo identificar el archivo", 404

    ruta = os.path.join(UPLOAD_FOLDER, filename)

    resultado = preparar_datos(ruta)
    df = resultado["dataframe"]

    fila_df = df[df["_excel_row"] == fila]

    if fila_df.empty:
        return "Fila no encontrada", 404

    fila_data = fila_df.iloc[0].to_dict()

    inconsistencias = cargar_resultados()

  
    detalle_lt = None
    otras_reglas = []

    for inc in inconsistencias:
        if int(inc.get("fila", -1)) == int(fila):
            if "leadtime_real" in inc and inc["leadtime_real"] is not None:
                if detalle_lt is None:
                    detalle_lt = inc
            else:
                otras_reglas.append(inc)

    detalle = detalle_lt if detalle_lt is not None else {}

    
    detalle_lifecycle = None

    for inc in inconsistencias:
        if int(inc.get("fila", -1)) == int(fila):
            if "dias_transcurridos" in inc:
                detalle_lifecycle = inc


   
    return render_template(
        "ver_fila_error.html",
        fila=fila,
        datos=fila_data,
        detalle=detalle,
        detalle_lifecycle=detalle_lifecycle,
        otras_reglas=otras_reglas,
        archivo=filename
    )





# ===============================
# Exportar a excel
# ===============================




@app.route("/exportar_excel", methods=["GET","POST"])
def exportar_excel():

    
    data = request.get_json(silent=True) or {}
    imagenes = data.get("imagenes", [])


    inconsistencias = cargar_resultados()
    anomalias = cargar_resultados_ia()

    df_inc = pd.DataFrame(inconsistencias)
    df_ano = pd.DataFrame(anomalias)

    output = io.BytesIO()

    fill_error = PatternFill(
        start_color="FFC7CE",
        end_color="FFC7CE",
        fill_type="solid"
    )

    with pd.ExcelWriter(output, engine="openpyxl") as writer:

        # =========================
        #  HOJA INCONSISTENCIAS
        # =========================
        if not df_inc.empty:

            df_inc.to_excel(writer, sheet_name="Inconsistencias", index=False)

            ws = writer.sheets["Inconsistencias"]

            
            #  crear tabla estilo Excel
            num_filas = ws.max_row
            num_cols = ws.max_column

            from openpyxl.utils import get_column_letter
            last_col = get_column_letter(num_cols)

            tabla = Table(
                displayName="TablaInconsistencias",
                ref=f"A1:{last_col}{num_filas}"
            )

            style = TableStyleInfo(
                name="TableStyleMedium9", 
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )

            tabla.tableStyleInfo = style
            ws.add_table(tabla)


            for row_idx, inc in enumerate(inconsistencias, start=2):

                campo1 = inc.get("campo_error_1")
                campo2 = inc.get("campo_error_2")

                for col_idx, col_name in enumerate(df_inc.columns, start=1):

                    if col_name == campo1 or col_name == campo2:
                        ws.cell(row=row_idx, column=col_idx).fill = fill_error


        # =========================
        #  HOJA ANOMALIAS
        # =========================
        if not df_ano.empty:
            df_ano.to_excel(writer, sheet_name="Anomalias", index=False)


        # =========================
        #  HOJA RESUMEN
        # =========================
        resumen = {
            "Total inconsistencias": [len(inconsistencias)],
            "Total anomalias": [len(anomalias)]
        }

        df_resumen = pd.DataFrame(resumen)
        df_resumen.to_excel(writer, sheet_name="Resumen", index=False)


        # =========================
        #  HOJA SEVERIDAD (grafico)
        # =========================
        if inconsistencias:
            df_severidad = pd.DataFrame([
                {"Severidad": inc.get("severidad", "N/A")}
                for inc in inconsistencias
            ])

            df_severidad = df_severidad.value_counts().reset_index()
            df_severidad.columns = ["Severidad", "Cantidad"]

            df_severidad.to_excel(writer, sheet_name="Severidad", index=False)


        # =========================
        #  HOJA REGLAS (grafico)
        # =========================
        if inconsistencias:
            df_reglas = pd.DataFrame([
                {"Regla": inc.get("regla_id", "N/A")}
                for inc in inconsistencias
            ])

            df_reglas = df_reglas.value_counts().reset_index()
            df_reglas.columns = ["Regla", "Cantidad"]

            df_reglas.to_excel(writer, sheet_name="Reglas", index=False)


        # =========================
        #  HOJA OWNER (slicer)
        # =========================
        if inconsistencias:
            df_owner = pd.DataFrame([
                {"Owner": inc.get("owner", "SIN OWNER")}
                for inc in inconsistencias
            ])

            df_owner = df_owner.value_counts().reset_index()
            df_owner.columns = ["Owner", "Cantidad"]

            df_owner.to_excel(writer, sheet_name="Owner", index=False)

        
        if imagenes:

            ws_graficos = writer.book.create_sheet("Graficos")

            row_pos = 1

            for img_base64 in imagenes:

                img_data = base64.b64decode(img_base64.split(",")[1])

                with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                    tmp.write(img_data)
                    tmp_path = tmp.name

                img = ExcelImage(tmp_path)

                
                img.width = 800
                img.height = 400

                img.anchor = f"A{row_pos}"
                ws_graficos.add_image(img)

                row_pos += 25

    output.seek(0)

    return send_file(
        output,
        download_name="reporte_auditoria.xlsx",
        as_attachment=True
    )


@app.route("/historico")
def historico():

    historico = cargar_json(HISTORICO_FILE)

    return render_template(
        "historico.html",
        historico=historico
    )

# ===============================
# EJECUCIÓN
# ===============================

if __name__ == "__main__":
    app.run(debug=True)
